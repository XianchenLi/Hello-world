import streamlit as st
import pandas as pd
import numpy as np
import base64
from io import BytesIO
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import warnings

# 忽略警告
warnings.filterwarnings('ignore')

# 设置页面
st.set_page_config(
    page_title="Excel对比与标记工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 页面标题
st.title("📊 Excel对比与标记工具")
st.markdown("上传两个Excel文件进行对比，支持单个Sheet或所有同名Sheet对比")

# 初始化session状态
if 'file1' not in st.session_state:
    st.session_state.file1 = None
if 'file2' not in st.session_state:
    st.session_state.file2 = None
if 'selected_sheet' not in st.session_state:
    st.session_state.selected_sheet = None
if 'all_sheets' not in st.session_state:
    st.session_state.all_sheets = False
if 'sheet_names1' not in st.session_state:
    st.session_state.sheet_names1 = []
if 'sheet_names2' not in st.session_state:
    st.session_state.sheet_names2 = []
if 'marked_results' not in st.session_state:
    st.session_state.marked_results = {}
if 'sheet_key_columns' not in st.session_state:
    st.session_state.sheet_key_columns = {}

# 颜色定义 - 使用aRGB格式 (8位十六进制值)
UNCHANGED_COLOR = "FFD3D3D3"  # 灰色 - 不变
ADDED_COLOR = "FF90EE90"     # 浅绿色 - 新增
DELETED_COLOR = "FFADD8E6"   # 浅蓝色 - 删除
MODIFIED_COLOR = "FFFFFF00"  # 黄色 - 修改单元格
HEADER_COLOR = "FFDAE8FC"    # 浅蓝色 - 表头

# 创建边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 创建两列布局
col1, col2 = st.columns(2)

# 上传文件1
with col1:
    st.subheader("上传原始文件")
    uploaded_file1 = st.file_uploader(
        "选择原始文件 (Excel)",
        type=["xlsx", "xls"],
        key="file1_uploader"
    )
    
    if uploaded_file1:
        st.session_state.file1 = uploaded_file1
        # 获取所有sheet名称
        try:
            excel_file = pd.ExcelFile(uploaded_file1)
            st.session_state.sheet_names1 = excel_file.sheet_names
            st.success(f"已上传: {uploaded_file1.name} ({len(st.session_state.sheet_names1)}个sheet)")
        except Exception as e:
            st.error(f"读取Excel文件出错: {str(e)}")

# 上传文件2
with col2:
    st.subheader("上传对比文件")
    uploaded_file2 = st.file_uploader(
        "选择对比文件 (Excel)",
        type=["xlsx", "xls"],
        key="file2_uploader"
    )
    
    if uploaded_file2:
        st.session_state.file2 = uploaded_file2
        # 获取所有sheet名称
        try:
            excel_file = pd.ExcelFile(uploaded_file2)
            st.session_state.sheet_names2 = excel_file.sheet_names
            st.success(f"已上传: {uploaded_file2.name} ({len(st.session_state.sheet_names2)}个sheet)")
        except Exception as e:
            st.error(f"读取Excel文件出错: {str(e)}")

# Excel文件处理函数
def read_excel(file, sheet_name=None):
    try:
        if sheet_name:
            return pd.read_excel(file, sheet_name=sheet_name)
        else:
            return pd.read_excel(file)
    except Exception as e:
        st.error(f"读取Excel文件出错: {str(e)}")
        return None

# 实际对比函数 - 在原始文件基础上标记修改
def compare_and_mark_changes(df1, df2, key_columns):
    # 检查是否有数据
    if df1 is None or df2 is None:
        return None, {}
    
    if df1.empty or df2.empty:
        return None, {}
    
    # 重置索引并添加原始索引列
    df1 = df1.reset_index(drop=True)
    df1['__original_index'] = df1.index
    df2 = df2.reset_index(drop=True)
    df2['__original_index'] = df2.index
    
    # 处理关键列
    if not key_columns:
        # 如果没有指定关键列，使用所有列
        key_columns = list(df1.columns)
    
    # 确保关键列在两个数据框中都存在
    valid_key_columns = [col for col in key_columns if col in df1.columns and col in df2.columns]
    if not valid_key_columns:
        # 使用默认索引
        valid_key_columns = ['__original_index']
    
    # 创建合并键
    df1['__merge_key'] = df1[valid_key_columns].astype(str).apply('|'.join, axis=1)
    df2['__merge_key'] = df2[valid_key_columns].astype(str).apply('|'.join, axis=1)
    
    # 找出新增行
    added = df2[~df2['__merge_key'].isin(df1['__merge_key'])].copy()
    
    # 找出删除行
    deleted = df1[~df1['__merge_key'].isin(df2['__merge_key'])].copy()
    
    # 找出共同行（可能被修改）
    common_keys = set(df1['__merge_key']) & set(df2['__merge_key'])
    common_df1 = df1[df1['__merge_key'].isin(common_keys)]
    common_df2 = df2[df2['__merge_key'].isin(common_keys)]
    
    # 创建标记后的数据框 - 基于原始文件
    marked_df = df1.copy()
    changes_dict = {}
    
    # 添加状态列
    marked_df['状态'] = '不变'
    
    # 标记删除行
    for idx in deleted.index:
        marked_df.at[idx, '状态'] = '删除'
    
    # 标记修改行
    for key in common_keys:
        row1 = common_df1[common_df1['__merge_key'] == key].iloc[0]
        row2 = common_df2[common_df2['__merge_key'] == key].iloc[0]
        idx = row1['__original_index']
        
        # 比较所有列
        cell_changes = {}
        has_changes = False
        
        for col in df1.columns:
            if col in ['__merge_key', '__original_index', '状态']:
                continue
                
            # 处理NaN值比较
            val1 = row1[col]
            val2 = row2[col]
            
            if pd.isna(val1) and pd.isna(val2):
                # 两个都是NaN，视为相同
                cell_changes[col] = False
            elif pd.isna(val1) or pd.isna(val2) or val1 != val2:
                # 值不同
                has_changes = True
                cell_changes[col] = True
                
                # 创建修改标记字符串
                old_val = str(val1) if not pd.isna(val1) else "空"
                new_val = str(val2) if not pd.isna(val2) else "空"
                marked_df.at[idx, col] = f"{old_val}->{new_val}"
            else:
                # 值相同
                cell_changes[col] = False
        
        if has_changes:
            marked_df.at[idx, '状态'] = '修改'
            changes_dict[idx] = cell_changes
    
    # 添加新增行
    for _, row in added.iterrows():
        new_row = row.copy()
        # 从新行中移除内部列
        if '__merge_key' in new_row:
            new_row.pop('__merge_key')
        if '__original_index' in new_row:
            new_row.pop('__original_index')
        
        # 设置状态
        new_row['状态'] = '新增'
        
        # 添加到标记数据框
        marked_df = pd.concat([marked_df, new_row.to_frame().T], ignore_index=True)
    
    # 移除内部列
    for col in ['__merge_key', '__original_index']:
        if col in marked_df.columns:
            marked_df.drop(columns=[col], inplace=True)
    
    # 重新排列列，将状态列放在最前面
    cols = marked_df.columns.tolist()
    if '状态' in cols:
        cols.remove('状态')
        cols.insert(0, '状态')
    marked_df = marked_df[cols]
    
    return marked_df, changes_dict

# 生成带标记的Excel文件
def generate_marked_excel(marked_results):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认创建的sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 为每个sheet创建标记结果
    for sheet_name, result in marked_results.items():
        marked_df = result['marked_df']
        changes_dict = result['changes_dict']
        
        # 创建sheet
        ws = wb.create_sheet(title=sheet_name[:31])  # 限制sheet名称长度
        
        # 写入标题行
        columns = marked_df.columns.tolist()
        
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据行
        for row_idx, (idx, row) in enumerate(marked_df.iterrows(), 2):
            status = row['状态']  # 获取状态
            
            # 写入每列数据
            for col_idx, col in enumerate(columns, 1):
                value = row[col]
                
                # 处理NaN值
                if pd.isna(value):
                    value = ""
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # 设置行状态背景色
                if status == '不变':
                    cell.fill = PatternFill(start_color=UNCHANGED_COLOR, end_color=UNCHANGED_COLOR, fill_type="solid")
                elif status == '新增':
                    cell.fill = PatternFill(start_color=ADDED_COLOR, end_color=ADDED_COLOR, fill_type="solid")
                elif status == '删除':
                    cell.fill = PatternFill(start_color=DELETED_COLOR, end_color=DELETED_COLOR, fill_type="solid")
                
                # 如果是修改行，并且有单元格被修改
                if status == '修改' and col != '状态':
                    # 检查这个单元格是否有修改
                    if idx in changes_dict and col in changes_dict[idx] and changes_dict[idx][col]:
                        cell.fill = PatternFill(start_color=MODIFIED_COLOR, end_color=MODIFIED_COLOR, fill_type="solid")
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            
            # 第一行是标题
            max_length = max(max_length, len(str(col[0].value)))
            
            for cell in col[1:]:
                try:
                    value = cell.value
                    if value is None:
                        continue
                    if isinstance(value, (int, float)):
                        value = str(value)
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大列宽
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# 获取sheet的列名
def get_sheet_columns(file, sheet_name):
    try:
        df = read_excel(file, sheet_name)
        if df is not None:
            return list(df.columns)
    except Exception as e:
        st.error(f"读取Sheet '{sheet_name}' 列名出错: {str(e)}")
    return []

# 对比选项
st.divider()
st.subheader("对比选项")

# 选择对比模式
col1, col2 = st.columns(2)
with col1:
    st.session_state.all_sheets = st.checkbox("对比所有同名Sheet", value=False)

with col2:
    if not st.session_state.all_sheets and st.session_state.sheet_names1 and st.session_state.sheet_names2:
        # 找出两个文件都有的sheet
        common_sheets = list(set(st.session_state.sheet_names1) & set(st.session_state.sheet_names2))
        
        if not common_sheets:
            st.warning("两个文件没有共同的Sheet名称")
        else:
            st.session_state.selected_sheet = st.selectbox(
                "选择要对比的Sheet",
                common_sheets,
                index=0
            )

# 关键列设置
if st.session_state.file1 and st.session_state.file2:
    # 单个Sheet模式下的关键列设置
    if not st.session_state.all_sheets and st.session_state.selected_sheet:
        st.subheader(f"关键列设置: {st.session_state.selected_sheet}")
        
        # 获取两个sheet的列名
        cols1 = get_sheet_columns(st.session_state.file1, st.session_state.selected_sheet)
        cols2 = get_sheet_columns(st.session_state.file2, st.session_state.selected_sheet)
        
        common_columns = list(set(cols1) & set(cols2))
        
        if common_columns:
            # 初始化当前sheet的关键列
            if st.session_state.selected_sheet not in st.session_state.sheet_key_columns:
                st.session_state.sheet_key_columns[st.session_state.selected_sheet] = []
            
            # 创建关键列选择器
            selected_keys = st.multiselect(
                f"选择用于比较的关键列 (可选)",
                common_columns,
                default=st.session_state.sheet_key_columns[st.session_state.selected_sheet],
                key=f"key_columns_{st.session_state.selected_sheet}"
            )
            
            # 更新关键列
            st.session_state.sheet_key_columns[st.session_state.selected_sheet] = selected_keys
        else:
            st.warning("两个Sheet没有共同的列名，无法设置关键列")
    
    # 所有Sheet模式下的关键列设置
    if st.session_state.all_sheets:
        st.subheader("关键列设置")
        common_sheets = list(set(st.session_state.sheet_names1) & set(st.session_state.sheet_names2))
        
        if common_sheets:
            for sheet_name in common_sheets:
                # 获取两个sheet的列名
                cols1 = get_sheet_columns(st.session_state.file1, sheet_name)
                cols2 = get_sheet_columns(st.session_state.file2, sheet_name)
                
                common_columns = list(set(cols1) & set(cols2))
                
                if common_columns:
                    # 初始化当前sheet的关键列
                    if sheet_name not in st.session_state.sheet_key_columns:
                        st.session_state.sheet_key_columns[sheet_name] = []
                    
                    # 创建关键列选择器
                    st.markdown(f"**Sheet: {sheet_name}**")
                    selected_keys = st.multiselect(
                        f"选择用于比较的关键列 (可选)",
                        common_columns,
                        default=st.session_state.sheet_key_columns[sheet_name],
                        key=f"key_columns_{sheet_name}"
                    )
                    
                    # 更新关键列
                    st.session_state.sheet_key_columns[sheet_name] = selected_keys
                else:
                    st.warning(f"Sheet '{sheet_name}' 没有共同的列名，无法设置关键列")
        else:
            st.warning("两个文件没有共同的Sheet名称")

# 对比按钮
if st.button("开始对比与标记", use_container_width=True, type="primary"):
    if st.session_state.file1 and st.session_state.file2:
        with st.spinner("正在对比文件并标记差异，请稍候..."):
            # 获取文件类型
            file1_type = st.session_state.file1.name.split('.')[-1].lower()
            file2_type = st.session_state.file2.name.split('.')[-1].lower()
            
            # 检查文件类型是否一致
            if file1_type != file2_type:
                st.error("错误：两个文件类型不一致，请上传相同类型的文件进行对比")
            else:
                # 确定要对比的sheet
                if st.session_state.all_sheets:
                    # 对比所有同名sheet
                    common_sheets = list(set(st.session_state.sheet_names1) & set(st.session_state.sheet_names2))
                    
                    if not common_sheets:
                        st.error("两个文件没有共同的Sheet名称")
                    else:
                        st.session_state.marked_results = {}
                        
                        for sheet_name in common_sheets:
                            try:
                                # 获取该sheet的关键列
                                key_columns = st.session_state.sheet_key_columns.get(sheet_name, [])
                                
                                # 读取两个sheet的数据
                                df1 = read_excel(st.session_state.file1, sheet_name)
                                df2 = read_excel(st.session_state.file2, sheet_name)
                                
                                # 对比并标记
                                marked_df, changes_dict = compare_and_mark_changes(df1, df2, key_columns)
                                
                                if marked_df is not None:
                                    st.session_state.marked_results[sheet_name] = {
                                        'marked_df': marked_df,
                                        'changes_dict': changes_dict,
                                        'key_columns': key_columns
                                    }
                            except Exception as e:
                                st.error(f"处理Sheet '{sheet_name}' 时出错: {str(e)}")
                        
                        if st.session_state.marked_results:
                            st.success(f"成功对比 {len(st.session_state.marked_results)} 个Sheet!")
                        else:
                            st.warning("没有生成任何对比结果")
                else:
                    # 对比单个sheet
                    if st.session_state.selected_sheet:
                        try:
                            # 获取该sheet的关键列
                            key_columns = st.session_state.sheet_key_columns.get(st.session_state.selected_sheet, [])
                            
                            # 读取两个sheet的数据
                            df1 = read_excel(st.session_state.file1, st.session_state.selected_sheet)
                            df2 = read_excel(st.session_state.file2, st.session_state.selected_sheet)
                            
                            # 对比并标记
                            marked_df, changes_dict = compare_and_mark_changes(df1, df2, key_columns)
                            
                            if marked_df is not None:
                                st.session_state.marked_results = {
                                    st.session_state.selected_sheet: {
                                        'marked_df': marked_df,
                                        'changes_dict': changes_dict,
                                        'key_columns': key_columns
                                    }
                                }
                                st.success(f"成功对比Sheet '{st.session_state.selected_sheet}'!")
                            else:
                                st.warning("没有生成对比结果")
                        except Exception as e:
                            st.error(f"处理Sheet '{st.session_state.selected_sheet}' 时出错: {str(e)}")
                    else:
                        st.warning("请选择要对比的Sheet")
    else:
        st.warning("请先上传两个Excel文件")

# 显示标记结果
if st.session_state.marked_results:
    st.divider()
    st.subheader("对比结果")
    
    # 显示sheet列表
    sheet_list = list(st.session_state.marked_results.keys())
    selected_sheet = st.selectbox("选择查看的Sheet", sheet_list, key="result_sheet_selector")
    
    if selected_sheet in st.session_state.marked_results:
        result = st.session_state.marked_results[selected_sheet]
        marked_df = result['marked_df']
        key_columns = result.get('key_columns', [])
        
        # 显示关键信息
        st.info(f"Sheet: **{selected_sheet}**")
        
        if key_columns:
            st.info(f"关键列: **{', '.join(key_columns)}**")
        else:
            st.info("关键列: 未设置 (使用行索引进行比较)")
        
        # 显示数据
        #st.dataframe(marked_df.head(10))
        
        # 显示统计信息
        status_counts = marked_df['状态'].value_counts()
        st.markdown("**状态统计:**")
        for status, count in status_counts.items():
            st.write(f"- {status}: {count}行")
        
        # 显示修改详情
        #if any(status == '修改' for status in marked_df['状态']):
            #st.markdown("**修改详情:**")
            #modified_rows = marked_df[marked_df['状态'] == '修改']
            #for idx, row in modified_rows.head(3).iterrows():
                #st.write(f"行 {idx+1}:")
                #for col in modified_rows.columns:
                    #if col != '状态' and '->' in str(row[col]):
                        #st.write(f"  - {col}: {row[col]}")
    
    st.divider()
    
    # 下载标记结果
    st.markdown("### 💾 下载对比结果")
    
    # 生成带标记的Excel文件
    try:
        marked_excel = generate_marked_excel(st.session_state.marked_results)
        
        # 创建下载按钮
        file_name = "对比结果_"
        if len(st.session_state.marked_results) == 1:
            file_name += f"{list(st.session_state.marked_results.keys())[0]}.xlsx"
        else:
            file_name += "多Sheet.xlsx"
        
        st.download_button(
            label="下载标记的Excel文件",
            data=marked_excel,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"生成标记的Excel文件时出错: {str(e)}")
        st.error("请确保上传的文件格式正确且包含有效数据")

# 使用说明
st.sidebar.title("使用说明")
st.sidebar.markdown("""
1. **上传文件**:
   - 左侧上传原始Excel文件
   - 右侧上传要对比的Excel文件

2. **选择对比模式**:
   - **单个Sheet**: 选择要对比的具体Sheet
   - **所有同名Sheet**: 自动对比两个文件中名称相同的所有Sheet

3. **设置关键列**:
   - 为每个Sheet单独设置用于比较的关键列
   - 关键列用于识别相同的行（如ID列）
   - 如果不设置关键列，将使用行索引进行比较

4. **执行对比**:
   - 点击"开始对比与标记"按钮

5. **查看结果**:
   - 预览标记后的数据
   - 查看状态统计信息
   - 查看修改详情

6. **下载结果**:
   - 下载带标记的Excel文件

**标记说明**:
- **不变**: 灰色背景 - 行在两个文件中完全相同
- **新增**: 绿色背景 - 行只存在于新文件中
- **删除**: 蓝色背景 - 行只存在于原始文件中
- **修改**: 黄色背景 - 单元格内容被修改

**修改单元格**:
- 显示格式: "原内容->修改后内容"
- 背景色: 黄色
- 示例: "张三->李四"

**注意事项**:
- 关键列应在两个文件中都存在
- 对于大型文件，对比可能需要一些时间
- 确保两个文件有相同的结构
- 所有同名Sheet模式只对比两个文件中都存在的Sheet
""", unsafe_allow_html=True)

# 添加页脚
st.divider()
st.caption("© 2023 Excel对比与标记工具 | 开发: Streamlit/Xianchen Li | 版本: 6.0")